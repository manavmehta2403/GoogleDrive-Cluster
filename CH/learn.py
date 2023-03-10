import openpyxl as op
from openpyxl.utils import column_index_from_string 
from openpyxl import Workbook
import pandas as pd

wb = op.load_workbook('Q3.xlsx')

#to print the number of sheets
print(wb.sheetnames)

#sheet number
sh = wb['Q.2']

print(sh.title)

cell = sh.cell(row=1, column =column_index_from_string('A'))
c = sh["C4"]

print(c.value)
print(c.coordinate)
print(cell.value)

#number of rows and columns
print(sh.max_row,sh.max_column)

##print("---------------------------------")
##for i in range (1,sh.max_row+1):
##    for j in range (1,sh.max_column+1):
##       c1 = sh.cell(row = i , column = j).value
##       print(c1, end= "\t")
##    print("\n")

###active sheet
##a_sh = wb.active #sheet number last working sheet
##
##print(a_sh.title)
w = Workbook()
w.create_sheet("Sheet")
sheet = w["Sheet"]

#read the excel file as data frame
df = pd.read_excel("Q3.xlsx","Q.1")

# Create a Pandas Excel writer using XlsxWriter as the engine.
wr = pd.ExcelWriter('demo.xlsx', engine='openpyxl')

# Convert the dataframe to an XlsxWriter Excel object.
df.to_excel(wr, sheet_name='Sheet', index = False)
wr.save()
w.save('demo.xlsx')

